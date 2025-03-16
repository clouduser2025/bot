from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import os
import datetime
import re
from telethon import TelegramClient
from telethon.sessions import StringSession
import asyncio
import logging
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
from contextlib import asynccontextmanager  # Added missing import

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="Newspaper Extraction API", version="1.0.0", description="API to extract text from newspaper PDFs via Telegram with progress tracking.")

# Telegram API Credentials
API_ID = 27889863
API_HASH = "df4d440af21594b001dc768518140c6b"
PHONE_NUMBER = "+919423780567"
SESSION_FILE = "toi_session_feb"
CHANNEL_USERNAME = "the_times_of_india_0"

# Directory to Save PDFs and Excel files
SAVE_DIR = "/root/automationbot/toi_editions"
os.makedirs(SAVE_DIR, exist_ok=True)
logger.info(f"SAVE_DIR set to: {SAVE_DIR}")

# Initialize Telegram Client
if os.path.exists(SESSION_FILE):
    with open(SESSION_FILE, "r") as f:
        session_str = f.read().strip()
    client = TelegramClient(StringSession(session_str), API_ID, API_HASH)
else:
    client = TelegramClient(StringSession(), API_ID, API_HASH)

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://tool.iysinfo.com", "http://localhost:8001"],  # Allow both production and local testing
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Input Model
class CityRequest(BaseModel):
    city: str

# Lifespan event handler
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Manage the lifecycle of the Telegram client."""
    try:
        await client.start(phone=PHONE_NUMBER)
        logger.info("Telegram client connected!")
        if not os.path.exists(SESSION_FILE):
            session_str = client.session.save()
            with open(SESSION_FILE, "w") as f:
                f.write(session_str)
            logger.info(f"Session saved to {SESSION_FILE}")
        yield
    except Exception as e:
        logger.error(f"Error in lifespan: {str(e)}", exc_info=True)
        raise
    finally:
        await client.disconnect()
        logger.info("Telegram client disconnected!")

app = FastAPI(title="Newspaper Extraction API", version="1.0.0", lifespan=lifespan, description="API to extract text from newspaper PDFs via Telegram with progress tracking.")

def append_to_excel(excel_path, data, first_write=False):
    """Append data to an Excel file incrementally."""
    try:
        if first_write:
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "Page No.", "Keyword", "City", "Extracted Text"])
            for row in data:
                ws.append([row["Date"], row["Page No."], row["Keyword"], row["City"], row["Extracted Text"]])
            wb.save(excel_path)
        else:
            wb = load_workbook(excel_path)
            ws = wb.active
            for row in data:
                ws.append([row["Date"], row["Page No."], row["Keyword"], row["City"], row["Extracted Text"]])
            wb.save(excel_path)
    except Exception as e:
        logger.error(f"Error appending to Excel file {excel_path}: {str(e)}")
        raise

def delete_previous_excel_files(current_date):
    """Delete Excel files from previous dates."""
    current_date_str = current_date.strftime('%d-%m-%Y')
    excel_pattern = re.compile(r"TOI_.*_(\d{2}-\d{2}-\d{4})_extracted\.xlsx$")
    for filename in os.listdir(SAVE_DIR):
        match = excel_pattern.match(filename)
        if match:
            file_date_str = match.group(1)
            if file_date_str != current_date_str:
                file_path = os.path.join(SAVE_DIR, filename)
                try:
                    os.remove(file_path)
                    logger.info(f"Deleted previous Excel file: {file_path}")
                except Exception as e:
                    logger.error(f"Error deleting Excel file {file_path}: {str(e)}")

def process_newspaper_task(city, today, background_tasks):
    """Background task to process newspaper extraction with tqdm progress."""
    simplified_pattern = rf"TOI.*{city}.*{today.strftime('%d-%m')}"
    file_path = None
    excel_path = os.path.join(SAVE_DIR, f"TOI_{city}_{today.strftime('%d-%m-%Y')}_extracted.xlsx")
    results = {"steps": [], "status": "completed", "extracted_results": [], "progress": {"current_page": 0, "total_pages": 0}}

    try:
        # Step 0: Clean up previous Excel files
        delete_previous_excel_files(today)

        # Step 1: Find PDF
        results["steps"].append({"step": "Finding PDF", "status": "started"})
        file_name = None
        message_to_download = None
        async def find_pdf():
            async for message in client.iter_messages(CHANNEL_USERNAME):
                if message.file and message.file.name:
                    if re.search(simplified_pattern, message.file.name, re.IGNORECASE):
                        nonlocal file_name, message_to_download
                        file_name = message.file.name
                        message_to_download = message
                        break
        asyncio.run(find_pdf())
        if file_name:
            results["steps"].append({"step": "Finding PDF", "status": "success", "filename": file_name})
        else:
            error_msg = f"No PDF found for {city} on {today.strftime('%d-%m-%Y')}"
            results["steps"].append({"step": "Finding PDF", "status": "failed", "error": error_msg})
            return results

        # Step 2: Download PDF
        results["steps"].append({"step": "Downloading PDF", "status": "started"})
        file_path = os.path.join(SAVE_DIR, file_name)
        asyncio.run(message_to_download.download_media(file=file_path))
        results["steps"].append({"step": "Downloading PDF", "status": "success", "file_path": file_path})

        # Step 3: Extract Text with tqdm
        results["steps"].append({"step": "Extracting Text", "status": "started"})
        keywords = ["Public Notice", "Tenders", "Property", "Plot", "Registry"]
        total_records = 0
        page_number = 1
        first_write = True

        with tqdm(total=100, desc="Extracting Pages", unit="page") as pbar:
            while True:
                image_path = os.path.join(SAVE_DIR, f"page_{page_number}.png")
                try:
                    images = convert_from_path(file_path, dpi=300, first_page=page_number, last_page=page_number)
                    if not images:
                        break
                    image = images[0]
                    image.save(image_path)
                    logger.info(f"Processing page {page_number}")
                    text = pytesseract.image_to_string(image)
                    paragraphs = text.split("\n\n")
                    page_results = []
                    for keyword in keywords:
                        for idx, paragraph in enumerate(paragraphs):
                            if keyword.lower() in paragraph.lower():
                                before = paragraphs[idx - 1] if idx > 0 else ""
                                after = paragraphs[idx + 1] if idx < len(paragraphs) - 1 else ""
                                full_section = f"{before}\n\n{paragraph}\n\n{after}".strip()
                                record = {
                                    "Date": today.strftime('%Y-%m-%d'),
                                    "Page No.": page_number,
                                    "Keyword": keyword,
                                    "City": city,
                                    "Extracted Text": full_section
                                }
                                page_results.append(record)
                                results["extracted_results"].append(record)

                    if page_results:
                        append_to_excel(excel_path, page_results, first_write)
                        total_records += len(page_results)
                        first_write = False

                    os.remove(image_path)
                    results["progress"]["current_page"] = page_number
                    results["progress"]["total_pages"] += 1
                    pbar.update(1)  # Update tqdm progress bar
                    page_number += 1
                except Exception as e:
                    logger.error(f"Error processing page {page_number}: {str(e)}")
                    break

        results["steps"].append({
            "step": "Extracting Text",
            "status": "success",
            "total_pages": results["progress"]["total_pages"],
            "records_extracted": total_records
        })
        results["pdf_path"] = file_path
        results["excel_path"] = excel_path

    except Exception as e:
        results["steps"].append({"step": "Processing", "status": "failed", "error": str(e)})
        results["status"] = "failed"
    finally:
        return results

@app.post("/process-newspaper/")
async def process_newspaper(city: CityRequest, background_tasks: BackgroundTasks):
    """Process newspaper extraction for a given city using POST request."""
    city = city.city.strip().title()
    today = datetime.datetime.today().date()

    # Trigger background task and return initial response
    background_tasks.add_task(process_newspaper_task, city, today, background_tasks)
    return {"status": "processing", "message": f"Started processing for {city}. Check back later for results.", "city": city, "date": today.strftime('%Y-%m-%d')}

@app.get("/get-progress/{city}/{date}")
async def get_progress(city: str, date: str):
    """Poll this endpoint to get the current progress and results."""
    excel_path = os.path.join(SAVE_DIR, f"TOI_{city}_{date}_extracted.xlsx")
    results = {
        "status": "processing",
        "progress": {"current_page": 0, "total_pages": 0},
        "extracted_results": []
    }
    if os.path.exists(excel_path):
        results["status"] = "completed"
        wb = load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            results["extracted_results"].append({
                "Date": row[0],
                "Page No.": row[1],
                "Keyword": row[2],
                "City": row[3],
                "Extracted Text": row[4]
            })
    return results

@app.get("/download-pdf/{city}/{date}")
async def download_pdf(city: str, date: str):
    """Download the processed PDF file."""
    pdf_path = os.path.join(SAVE_DIR, f"TOI_{city}_{date.split('-')[0]}-{date.split('-')[1]}.pdf")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail="PDF file not found")
    return FileResponse(pdf_path, filename=f"TOI_{city}_{date}.pdf")

@app.get("/download-excel/{city}/{date}")
async def download_excel(city: str, date: str):
    """Download the extracted data as an Excel file."""
    excel_path = os.path.join(SAVE_DIR, f"TOI_{city}_{date}_extracted.xlsx")
    if not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Excel file not found")
    return FileResponse(excel_path, filename=f"TOI_{city}_{date}_extracted.xlsx")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8005)  # Removed SSL parameters